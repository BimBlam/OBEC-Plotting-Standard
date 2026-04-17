"""
Tests for batteryplot.reader — multi-format ingestion and mass detection.
"""
from __future__ import annotations

import os
import textwrap
from pathlib import Path

import pandas as pd
import pytest

from batteryplot.config import default_config
from batteryplot.reader import (
    SUPPORTED_EXTENSIONS,
    _resolve_mass,
    _sniff_delimiter,
    is_supported,
    load_file,
    prompt_mass_if_default,
)

# ---------------------------------------------------------------------------
# Fixtures: minimal test files
# ---------------------------------------------------------------------------

ARBIN_CSV_CONTENT = textwrap.dedent("""\
    Today's Date ,4/14/2026
    Date of Test:,4/10/2026 3:02:23 AM
    Test Time,Cycle P,Current (A),Voltage (V),Cycle C,Capacity (AHr),Step Time,Energy (WHr),MD
      0d 00:00:00,1,0.000,3.800,1,0.000,  0d 00:00:00,0.000,R
      0d 00:00:10,1,0.100,3.810,1,0.000278,  0d 00:00:10,0.001060,C
      0d 00:00:20,1,-0.100,3.790,1,0.000278,  0d 00:00:10,0.001060,D
""")

# Real Maccor .txt format: SCap line carries active mass in grams (e.g. 0.023331 g).
# The old interpretation of SCap as Ah/g was incorrect.
MACCOR_TXT_CONTENT = textwrap.dedent("""\
    Maccor Series 4000
    Procedure:\tTest_001
    SCap:\t0.023331 g
    Rec#\tCyc#\tStep\tTest Time\tStep Time\tAmp-hr\tWatt-hr\tAmps\tVolts\tState\tES\tDPt Time
    1\t1\t1\t00:00:00.0\t00:00:00.0\t0\t0\t0\t3.8\tR\t0\t01/01/2026 00:00:00
    2\t1\t2\t00:00:10.0\t00:00:10.0\t0.000278\t0.00106\t0.1\t3.81\tC\t0\t01/01/2026 00:00:10
    3\t1\t3\t00:00:20.0\t00:00:10.0\t0.000556\t0.00213\t-0.1\t3.79\tD\t0\t01/01/2026 00:00:20
""")

# Variant where both a Mass: line and SCap: line are present.
# Mass: takes precedence over SCap: for active_mass_g_from_file.
MACCOR_TXT_BOTH_FIELDS = textwrap.dedent("""\
    Maccor Series 4000
    Mass:\t0.019
    SCap:\t0.023331 g
    Rec#\tCyc#\tStep\tTest Time\tAmp-hr\tAmps\tVolts\tState
    1\t1\t1\t00:00:00.0\t0\t0\t3.8\tR
    2\t1\t2\t00:00:10.0\t0.000278\t0.1\t3.81\tC
""")

# Default-mass variant via the SCap line (the common Maccor case: user never
# changed the mass, so SCap shows 1 g).
MACCOR_TXT_DEFAULT_MASS = textwrap.dedent("""\
    Maccor Series 4000
    SCap:\t1 g
    Rec#\tCyc#\tStep\tTest Time\tAmp-hr\tAmps\tVolts\tState
    1\t1\t1\t00:00:00.0\t0\t0\t3.8\tR
    2\t1\t2\t00:00:10.0\t0.000278\t0.1\t3.81\tC
""")


@pytest.fixture
def arbin_csv(tmp_path) -> Path:
    p = tmp_path / "arbin_cell.csv"
    p.write_text(ARBIN_CSV_CONTENT, encoding="utf-8")
    return p


@pytest.fixture
def maccor_txt(tmp_path) -> Path:
    p = tmp_path / "maccor_cell.txt"
    p.write_text(MACCOR_TXT_CONTENT, encoding="utf-8")
    return p


@pytest.fixture
def maccor_txt_default_mass(tmp_path) -> Path:
    p = tmp_path / "maccor_default_mass.txt"
    p.write_text(MACCOR_TXT_DEFAULT_MASS, encoding="utf-8")
    return p


@pytest.fixture
def maccor_txt_both_fields(tmp_path) -> Path:
    p = tmp_path / "maccor_both_fields.txt"
    p.write_text(MACCOR_TXT_BOTH_FIELDS, encoding="utf-8")
    return p


# ---------------------------------------------------------------------------
# is_supported / SUPPORTED_EXTENSIONS
# ---------------------------------------------------------------------------

def test_supported_extensions_include_csv():
    assert ".csv" in SUPPORTED_EXTENSIONS

def test_supported_extensions_include_txt():
    assert ".txt" in SUPPORTED_EXTENSIONS

def test_supported_extensions_include_xls():
    assert ".xls" in SUPPORTED_EXTENSIONS

def test_supported_extensions_include_xlsx():
    assert ".xlsx" in SUPPORTED_EXTENSIONS

def test_is_supported_csv(tmp_path):
    p = tmp_path / "test.csv"
    p.touch()
    assert is_supported(p)

def test_is_supported_txt(tmp_path):
    p = tmp_path / "test.txt"
    p.touch()
    assert is_supported(p)

def test_is_supported_unsupported(tmp_path):
    p = tmp_path / "test.docx"
    p.touch()
    assert not is_supported(p)

def test_is_supported_case_insensitive(tmp_path):
    p = tmp_path / "test.CSV"
    p.touch()
    assert is_supported(p)


# ---------------------------------------------------------------------------
# Delimiter sniffing
# ---------------------------------------------------------------------------

def test_sniff_delimiter_comma(arbin_csv):
    assert _sniff_delimiter(arbin_csv) == ","

def test_sniff_delimiter_tab(maccor_txt):
    assert _sniff_delimiter(maccor_txt) == "\t"


# ---------------------------------------------------------------------------
# load_file — CSV
# ---------------------------------------------------------------------------

def test_load_file_csv_returns_dataframe(arbin_csv):
    cfg = default_config()
    raw_df, metadata, mass_g = load_file(arbin_csv, cfg)
    assert isinstance(raw_df, pd.DataFrame)
    assert len(raw_df) >= 2

def test_load_file_csv_has_expected_columns(arbin_csv):
    cfg = default_config()
    raw_df, _, _ = load_file(arbin_csv, cfg)
    cols_lower = [c.lower() for c in raw_df.columns]
    assert any("current" in c for c in cols_lower)
    assert any("voltage" in c for c in cols_lower)

def test_load_file_csv_metadata_extracted(arbin_csv):
    cfg = default_config()
    _, metadata, _ = load_file(arbin_csv, cfg)
    assert isinstance(metadata, dict)
    # Should have captured "Today's Date" or similar
    assert len(metadata) >= 1

def test_load_file_csv_no_mass_in_arbin(arbin_csv):
    cfg = default_config()
    _, _, mass_g = load_file(arbin_csv, cfg)
    # Arbin CSV has no embedded mass → None
    assert mass_g is None


# ---------------------------------------------------------------------------
# load_file — TXT (Maccor-style)
# ---------------------------------------------------------------------------

def test_load_file_txt_returns_dataframe(maccor_txt):
    cfg = default_config()
    raw_df, metadata, mass_g = load_file(maccor_txt, cfg)
    assert isinstance(raw_df, pd.DataFrame)
    assert len(raw_df) >= 2

def test_load_file_txt_extracts_scap_as_mass(maccor_txt):
    """SCap line in Maccor .txt is the active mass in grams (0.023331 g here)."""
    cfg = default_config()
    _, metadata, mass_g = load_file(maccor_txt, cfg)
    # SCap: 0.023331 g  →  active_mass_g_from_file = 0.023331
    assert mass_g == pytest.approx(0.023331, rel=1e-4)
    assert "SCap" in metadata
    assert float(metadata["SCap"]) == pytest.approx(0.023331, rel=1e-4)

def test_load_file_txt_scap_stored_in_active_mass(maccor_txt):
    """active_mass_g_from_file is populated from the SCap line."""
    cfg = default_config()
    _, metadata, _ = load_file(maccor_txt, cfg)
    assert "active_mass_g_from_file" in metadata
    assert float(metadata["active_mass_g_from_file"]) == pytest.approx(0.023331, rel=1e-4)

def test_load_file_txt_mass_row_overrides_scap(maccor_txt_both_fields):
    """Explicit Mass: row takes precedence over SCap: row."""
    cfg = default_config()
    _, metadata, mass_g = load_file(maccor_txt_both_fields, cfg)
    # Mass: 0.019 should win over SCap: 0.023331
    assert mass_g == pytest.approx(0.019, rel=1e-3)
    # SCap should still be recorded as metadata
    assert "SCap" in metadata

def test_load_file_txt_mass_extracted_with_sentinel_warning(maccor_txt):
    """
    0.023331 g is below the 1 g sentinel, so the warning fires.  This is
    expected for real coin-cell masses; the sentinel exists only to catch the
    1 g Maccor factory default.
    """
    cfg = default_config()
    _, metadata, mass_g = load_file(maccor_txt, cfg)
    assert mass_g is not None
    assert mass_g == pytest.approx(0.023331, rel=1e-4)
    # 0.023331 < 1.0 sentinel → warning is expected
    assert "mass_warning" in metadata

def test_load_file_txt_config_mass_overrides_file(maccor_txt):
    """When config.active_mass_g is set explicitly it overrides the file value."""
    cfg = default_config()
    object.__setattr__(cfg, "active_mass_g", 0.019)
    _, metadata, mass_g = load_file(maccor_txt, cfg)
    assert mass_g == pytest.approx(0.019)
    assert mass_g is not None


def test_load_file_txt_scap_default_mass_triggers_warning(maccor_txt_default_mass):
    """SCap: 1 g in the file → mass_warning because 1 g is the sentinel default."""
    cfg = default_config()
    _, metadata, mass_g = load_file(maccor_txt_default_mass, cfg)
    assert mass_g == pytest.approx(1.0)
    assert "mass_warning" in metadata
    assert "default" in metadata["mass_warning"].lower()

def test_load_file_txt_has_expected_columns(maccor_txt):
    cfg = default_config()
    raw_df, _, _ = load_file(maccor_txt, cfg)
    cols_lower = [c.lower() for c in raw_df.columns]
    # Maccor uses "Amps" and "Volts"
    assert any("amp" in c or "current" in c for c in cols_lower)
    assert any("volt" in c for c in cols_lower)


# ---------------------------------------------------------------------------
# Mass resolution logic
# ---------------------------------------------------------------------------

def test_resolve_mass_config_overrides_file():
    """config.active_mass_g takes precedence over file-embedded value."""
    cfg = default_config()
    object.__setattr__(cfg, "active_mass_g", 0.015)
    metadata = {"active_mass_g_from_file": "0.020"}
    mass = _resolve_mass(metadata, cfg)
    assert mass == pytest.approx(0.015)

def test_resolve_mass_config_default_flagged():
    """Config mass of 1.0 g triggers the warning."""
    cfg = default_config()
    object.__setattr__(cfg, "active_mass_g", 1.0)
    metadata: dict = {}
    mass = _resolve_mass(metadata, cfg)
    assert mass == pytest.approx(1.0)
    assert "mass_warning" in metadata

def test_resolve_mass_file_value_used_when_config_none():
    """File-embedded mass used when config.active_mass_g is None."""
    cfg = default_config()
    metadata = {"active_mass_g_from_file": "0.019"}
    mass = _resolve_mass(metadata, cfg)
    assert mass == pytest.approx(0.019)

def test_resolve_mass_file_default_flagged():
    """File-embedded mass of 1.0 g is flagged as default."""
    cfg = default_config()
    metadata = {"active_mass_g_from_file": "1.0"}
    mass = _resolve_mass(metadata, cfg)
    assert mass == pytest.approx(1.0)
    assert "mass_warning" in metadata

def test_resolve_mass_returns_none_when_absent():
    cfg = default_config()
    metadata: dict = {}
    mass = _resolve_mass(metadata, cfg)
    assert mass is None

def test_resolve_mass_custom_threshold():
    """A custom threshold of 2.0 g flags a 1.5 g mass as default."""
    cfg = default_config()
    object.__setattr__(cfg, "default_mass_threshold_g", 2.0)
    metadata = {"active_mass_g_from_file": "1.5"}
    mass = _resolve_mass(metadata, cfg)
    assert mass == pytest.approx(1.5)
    assert "mass_warning" in metadata


# ---------------------------------------------------------------------------
# discover_input_files
# ---------------------------------------------------------------------------

def test_discover_input_files_finds_csv(tmp_path):
    from batteryplot.io import discover_input_files
    (tmp_path / "cell_a.csv").touch()
    (tmp_path / "cell_b.CSV").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 2

def test_discover_input_files_finds_txt(tmp_path):
    from batteryplot.io import discover_input_files
    (tmp_path / "cell_a.txt").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 1

def test_discover_input_files_finds_mixed(tmp_path):
    from batteryplot.io import discover_input_files
    (tmp_path / "a.csv").touch()
    (tmp_path / "b.txt").touch()
    (tmp_path / "c.xls").touch()
    (tmp_path / "d.xlsx").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 4

def test_discover_input_files_ignores_hidden(tmp_path):
    from batteryplot.io import discover_input_files
    (tmp_path / ".hidden.csv").touch()
    (tmp_path / "~lock.csv").touch()
    (tmp_path / "real.csv").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 1
    assert found[0].name == "real.csv"

def test_discover_input_files_ignores_subdir(tmp_path):
    from batteryplot.io import discover_input_files
    subdir = tmp_path / "output" / "cell_001" / "data"
    subdir.mkdir(parents=True)
    (subdir / "processed.csv").write_text("col1,col2\n1,2\n")
    (tmp_path / "real.csv").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 1
    assert found[0].name == "real.csv"

def test_discover_input_files_ignores_batch_summary(tmp_path):
    from batteryplot.io import discover_input_files
    (tmp_path / "batch_summary.csv").touch()
    (tmp_path / "batch_summary.xlsx").touch()
    (tmp_path / "real_cell.csv").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 1
    assert found[0].name == "real_cell.csv"

def test_discover_input_files_ignores_unsupported(tmp_path):
    from batteryplot.io import discover_input_files
    (tmp_path / "notes.docx").touch()
    (tmp_path / "image.png").touch()
    (tmp_path / "cell.csv").touch()
    found = discover_input_files(tmp_path)
    assert len(found) == 1

def test_discover_input_files_sorted(tmp_path):
    from batteryplot.io import discover_input_files
    for name in ["cell_c.csv", "cell_a.csv", "cell_b.csv"]:
        (tmp_path / name).touch()
    found = discover_input_files(tmp_path)
    assert [p.name for p in found] == ["cell_a.csv", "cell_b.csv", "cell_c.csv"]

def test_discover_input_files_missing_dir():
    from batteryplot.io import discover_input_files
    with pytest.raises(FileNotFoundError):
        discover_input_files(Path("/nonexistent/path/xyz"))

def test_discover_backwards_compat_alias(tmp_path):
    """discover_csv_files should still work as an alias."""
    from batteryplot.io import discover_csv_files
    (tmp_path / "cell.csv").touch()
    found = discover_csv_files(tmp_path)
    assert len(found) == 1


# ---------------------------------------------------------------------------
# prompt_mass_if_default
# ---------------------------------------------------------------------------

def test_prompt_mass_if_default_passthrough_none(tmp_path):
    """None mass passes through unchanged."""
    cfg = default_config()
    p = tmp_path / "cell.txt"
    result = prompt_mass_if_default(p, None, cfg)
    assert result is None

def test_prompt_mass_if_default_passthrough_real_mass(tmp_path):
    """A real (non-sentinel) mass passes through unchanged."""
    cfg = default_config()
    p = tmp_path / "cell.txt"
    result = prompt_mass_if_default(p, 0.023, cfg)
    assert result == pytest.approx(0.023)

def test_prompt_mass_if_default_noninteractive_skips_prompt(tmp_path, monkeypatch):
    """
    When stdin is not a tty (non-interactive / batch mode), the function
    returns the sentinel unchanged without prompting.
    """
    cfg = default_config()
    p = tmp_path / "cell.txt"
    # Simulate non-interactive stdin
    monkeypatch.setattr("sys.stdin", open(os.devnull, "r"))
    result = prompt_mass_if_default(p, 1.0, cfg)
    assert result == pytest.approx(1.0)

def test_prompt_mass_if_default_interactive_accepts_new_value(tmp_path, monkeypatch):
    """When stdin is a tty and the user enters a valid mass, it is returned."""
    import io as _io
    cfg = default_config()
    p = tmp_path / "cell.txt"
    # Simulate tty + user enters "0.014"
    monkeypatch.setattr("sys.stdin.isatty", lambda: True)
    monkeypatch.setattr("builtins.input", lambda prompt="": "0.014")
    result = prompt_mass_if_default(p, 1.0, cfg)
    assert result == pytest.approx(0.014)

def test_prompt_mass_if_default_interactive_blank_keeps_default(tmp_path, monkeypatch):
    """When the user presses Enter (blank), the sentinel mass is kept."""
    cfg = default_config()
    p = tmp_path / "cell.txt"
    monkeypatch.setattr("sys.stdin.isatty", lambda: True)
    monkeypatch.setattr("builtins.input", lambda prompt="": "")
    result = prompt_mass_if_default(p, 1.0, cfg)
    assert result == pytest.approx(1.0)

def test_prompt_mass_if_default_interactive_invalid_input_keeps_default(tmp_path, monkeypatch):
    """When the user enters something non-numeric, the sentinel is kept."""
    cfg = default_config()
    p = tmp_path / "cell.txt"
    monkeypatch.setattr("sys.stdin.isatty", lambda: True)
    monkeypatch.setattr("builtins.input", lambda prompt="": "not_a_number")
    result = prompt_mass_if_default(p, 1.0, cfg)
    assert result == pytest.approx(1.0)

def test_prompt_mass_if_default_interactive_negative_keeps_default(tmp_path, monkeypatch):
    """A non-positive entered value is rejected and the sentinel is kept."""
    cfg = default_config()
    p = tmp_path / "cell.txt"
    monkeypatch.setattr("sys.stdin.isatty", lambda: True)
    monkeypatch.setattr("builtins.input", lambda prompt="": "-0.5")
    result = prompt_mass_if_default(p, 1.0, cfg)
    assert result == pytest.approx(1.0)


# ---------------------------------------------------------------------------
# Excel J1/K1 mass extraction
# ---------------------------------------------------------------------------

def _make_minimal_excel(tmp_path: Path, j1_val: str, k1_val) -> Path:
    """
    Build a minimal .xlsx file where:
    - Row 0 (1-indexed row 1): J1 = j1_val (col 9), K1 = k1_val (col 10)
    - Row 1: column headers (Rec#, Cyc#, ..., Amps, Volts, State)
    - Row 2+: numeric data
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    # Fill A1..I1 with empty so column indices are correct
    for col in range(1, 10):  # cols 1-9 = A-I
        ws.cell(row=1, column=col, value="")
    ws.cell(row=1, column=10, value=j1_val)   # J1
    ws.cell(row=1, column=11, value=k1_val)   # K1
    # Header row (row 2 = index 1)
    headers = ["Rec#", "Cyc#", "Step", "Test Time", "Amp-hr", "Watt-hr", "Amps", "Volts", "State"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=2, column=ci, value=h)
    # Data row
    ws.cell(row=3, column=1, value=1)
    ws.cell(row=3, column=7, value=0.1)   # Amps
    ws.cell(row=3, column=8, value=3.81)  # Volts
    ws.cell(row=3, column=9, value="C")   # State
    p = tmp_path / "maccor_test.xlsx"
    wb.save(str(p))
    return p


def test_excel_j1_scap_k1_mass_extracted(tmp_path):
    """J1='SCap', K1=0.023331 → active_mass_g_from_file = 0.023331."""
    p = _make_minimal_excel(tmp_path, "SCap", 0.023331)
    cfg = default_config()
    _, metadata, mass_g = load_file(p, cfg)
    assert mass_g == pytest.approx(0.023331, rel=1e-4)
    assert metadata.get("excel_J1") == "SCap"

def test_excel_j1_scap_k1_default_triggers_warning(tmp_path):
    """J1='SCap', K1=1 (factory default) → mass_warning is set."""
    p = _make_minimal_excel(tmp_path, "SCap", 1)
    cfg = default_config()
    _, metadata, mass_g = load_file(p, cfg)
    assert mass_g == pytest.approx(1.0)
    assert "mass_warning" in metadata

def test_excel_j1_not_scap_no_mass(tmp_path):
    """J1='SomeOtherLabel', K1=0.023 → no active_mass_g_from_file extracted via J1/K1."""
    p = _make_minimal_excel(tmp_path, "Voltage", 0.023)
    cfg = default_config()
    _, metadata, mass_g = load_file(p, cfg)
    # "Voltage" in J1 does not match SCap or mass/wt patterns
    assert metadata.get("active_mass_g_from_file") is None
    assert mass_g is None
