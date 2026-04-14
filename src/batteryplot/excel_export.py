"""
batteryplot.excel_export
========================
Export processed battery data to a multi-sheet Excel workbook.

Sheet layout:
1. column_map          — raw ↔ canonical column mapping
2. metadata            — header key/value pairs extracted from the file
3. cleaned_timeseries  — canonical-column analysis DataFrame (truncated if > Excel row limit)
4. cycle_summary       — per-cycle statistics
5. pulse_events        — detected pulse events (if any)
6. derived_metrics     — any computed metrics not in other sheets
7. plot_availability   — which plots were generated vs. placeholder

All sheets have:
- Bold, light-blue header row
- Freeze pane at A2 (top row frozen)
- Auto-adjusted column widths (capped at 40 characters)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Optional

import pandas as pd

logger = logging.getLogger("batteryplot")

# Excel maximum rows per sheet (including the header row)
_EXCEL_MAX_ROWS = 1_048_575  # One less than 1,048,576 to leave room for header

# openpyxl formatting constants
_HEADER_FILL_COLOR = "BDD7EE"   # Light blue (Excel default table header colour)
_HEADER_FONT_BOLD = True


def _apply_header_style(ws) -> None:
    """
    Bold + light-blue-fill the top row of an openpyxl worksheet, and freeze
    the pane at A2 so the header stays visible when scrolling.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet to format.
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL_COLOR)
    font = Font(bold=True)

    for cell in ws[1]:  # first row
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.freeze_panes = "A2"


def _auto_column_widths(ws, max_width: int = 40) -> None:
    """
    Set column widths based on the maximum content length in each column,
    capped at *max_width* characters.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Worksheet to resize.
    max_width : int
        Maximum column width in characters.
    """
    for col in ws.columns:
        col_letter = col[0].column_letter
        best_width = 0
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                best_width = max(best_width, length)
        ws.column_dimensions[col_letter].width = min(best_width + 2, max_width)


def _write_sheet(
    writer: "pd.ExcelWriter",
    df: pd.DataFrame,
    sheet_name: str,
    note: Optional[str] = None,
) -> None:
    """
    Write *df* to *sheet_name* in *writer*, then apply header formatting.

    If *note* is given it is placed in cell A1 before the data (the data
    is shifted down by one row).

    Parameters
    ----------
    writer : pd.ExcelWriter
        Open ExcelWriter bound to an openpyxl workbook.
    df : pd.DataFrame
        DataFrame to write.
    sheet_name : str
        Target sheet name (max 31 characters; will be truncated if longer).
    note : str, optional
        Prepend a note row above the header.
    """
    sheet_name = sheet_name[:31]  # Excel sheet name limit

    if note:
        # Write note first, then data starting at row 2
        note_df = pd.DataFrame([[note] + [""] * (len(df.columns) - 1)], columns=df.columns)
        combined = pd.concat([note_df, df], ignore_index=True)
        combined.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    ws = writer.sheets[sheet_name]
    _apply_header_style(ws)
    _auto_column_widths(ws)


def export_excel(
    output_path: Path,
    raw_df: pd.DataFrame,
    column_map: Dict[str, str],
    analysis_df: pd.DataFrame,
    cycle_summary: pd.DataFrame,
    plot_availability: pd.DataFrame,
    metadata: Dict[str, str],
    pulse_df: Optional[pd.DataFrame] = None,
) -> Path:
    """
    Write a multi-sheet Excel workbook for a single processed cell.

    Sheets written
    --------------
    1. **column_map**         — two columns: raw_column_name, canonical_name.
                                Unmapped raw columns are appended with an empty canonical.
    2. **metadata**           — key/value pairs from the file header.
    3. **cleaned_timeseries** — analysis_df (canonical columns, numeric values).
                                Truncated to 1,048,575 rows; a note is inserted in
                                cell A1 if truncation occurred.
    4. **cycle_summary**      — cycle-level summary stats.
    5. **pulse_events**       — pulse_df if not None and not empty.
    6. **derived_metrics**    — columns computed by the pipeline that are not in
                                the other sheets (c_rate, specific_capacity_mah_g, etc.).
    7. **plot_availability**  — which plots were generated vs. placeholder.

    Parameters
    ----------
    output_path : Path
        Destination ``.xlsx`` file.
    raw_df : pd.DataFrame
        Raw DataFrame with original column names.
    column_map : Dict[str, str]
        Mapping ``{raw_column → canonical_column}`` for matched columns.
    analysis_df : pd.DataFrame
        Canonical analysis DataFrame produced by :func:`build_analysis_df`.
    cycle_summary : pd.DataFrame
        Per-cycle summary produced by :func:`compute_cycle_summary`.
    plot_availability : pd.DataFrame
        Availability table from :func:`build_plot_availability`.
    metadata : Dict[str, str]
        Header key/value pairs extracted from the file.
    pulse_df : pd.DataFrame, optional
        Pulse events DataFrame; skipped if None or empty.

    Returns
    -------
    Path
        The *output_path* that was written.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Exporting Excel workbook: %s", output_path)

    # ------------------------------------------------------------------
    # Sheet 1: column_map
    # ------------------------------------------------------------------
    # Build rows: mapped + unmapped
    mapped_rows = [
        {"raw_column_name": raw, "canonical_name": canon}
        for raw, canon in column_map.items()
    ]
    unmapped_raw = [c for c in raw_df.columns if c not in column_map]
    unmapped_rows = [
        {"raw_column_name": raw, "canonical_name": ""}
        for raw in unmapped_raw
    ]
    colmap_df = pd.DataFrame(mapped_rows + unmapped_rows)
    if colmap_df.empty:
        colmap_df = pd.DataFrame(columns=["raw_column_name", "canonical_name"])

    # ------------------------------------------------------------------
    # Sheet 2: metadata
    # ------------------------------------------------------------------
    meta_df = pd.DataFrame(
        list(metadata.items()), columns=["key", "value"]
    )
    if meta_df.empty:
        meta_df = pd.DataFrame(columns=["key", "value"])

    # ------------------------------------------------------------------
    # Sheet 3: cleaned_timeseries — truncate if necessary
    # ------------------------------------------------------------------
    truncated = False
    ts_df = analysis_df.copy() if analysis_df is not None else pd.DataFrame()
    if len(ts_df) > _EXCEL_MAX_ROWS:
        truncated = True
        ts_note = (
            f"WARNING: Data truncated to {_EXCEL_MAX_ROWS:,} rows "
            f"(original: {len(ts_df):,} rows). "
            "Full data is available in cleaned_timeseries.csv."
        )
        ts_df = ts_df.iloc[:_EXCEL_MAX_ROWS]
        logger.warning(
            "cleaned_timeseries sheet truncated to %d rows (was %d).",
            _EXCEL_MAX_ROWS, len(analysis_df),
        )
    else:
        ts_note = None

    # ------------------------------------------------------------------
    # Sheet 6: derived_metrics — extract columns not in core timeseries
    # ------------------------------------------------------------------
    # These are columns added by transforms (c_rate, specific_capacity_mah_g, etc.)
    # that are not part of the original canonical mapping
    derived_cols = [
        col for col in (analysis_df.columns if analysis_df is not None else [])
        if col not in set(column_map.values())
        and col not in {"step_type_inferred", "segment_id"}
        and not col.startswith("_")
    ]
    if derived_cols and analysis_df is not None and not analysis_df.empty:
        id_cols = [c for c in ["elapsed_time_s", "cycle_index", "step_index"] if c in analysis_df.columns]
        keep_cols = list(dict.fromkeys(id_cols + derived_cols))  # deduplicated, ordered
        derived_df = analysis_df[keep_cols].copy()
        # Truncate to Excel limit
        if len(derived_df) > _EXCEL_MAX_ROWS:
            derived_df = derived_df.iloc[:_EXCEL_MAX_ROWS]
    else:
        derived_df = pd.DataFrame(columns=["no_derived_metrics"])

    # ------------------------------------------------------------------
    # Write all sheets
    # ------------------------------------------------------------------
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(writer, colmap_df, "column_map")
        _write_sheet(writer, meta_df, "metadata")

        if ts_note:
            _write_sheet(writer, ts_df, "cleaned_timeseries", note=ts_note)
        else:
            _write_sheet(writer, ts_df, "cleaned_timeseries")

        cs_df = cycle_summary if (cycle_summary is not None and not cycle_summary.empty) else pd.DataFrame(columns=["no_data"])
        _write_sheet(writer, cs_df, "cycle_summary")

        if pulse_df is not None and not pulse_df.empty:
            _write_sheet(writer, pulse_df, "pulse_events")

        _write_sheet(writer, derived_df, "derived_metrics")

        pa_df = plot_availability if (plot_availability is not None and not plot_availability.empty) else pd.DataFrame(columns=["no_data"])
        _write_sheet(writer, pa_df, "plot_availability")

    logger.info(
        "Excel workbook written: %s (sheets: column_map, metadata, "
        "cleaned_timeseries%s, cycle_summary, %spulse_events, derived_metrics, plot_availability)",
        output_path,
        " [truncated]" if truncated else "",
        "" if pulse_df is not None and not pulse_df.empty else "[no] ",
    )
    return output_path
